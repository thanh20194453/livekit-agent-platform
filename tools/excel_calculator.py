import shutil
from typing import Any, Callable, Dict, List, Literal, Optional, Set
from pydantic import BaseModel
import time
import os
import asyncio
import logging
import aiohttp


from utils.async_cache_file import AsyncStaticExcelCache
async_cache = AsyncStaticExcelCache(cache_dir="./temps/excel")

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None
    logger.warning("openpyxl is not installed. ExcelCalculator will not work without it.")


class ExcelRef(BaseModel):
    cell_ref: str
    value_type: Literal["number", "string", "date", "boolean", "percentage"]
    value: Any

class ExcelCalculator:
    def __init__(self, **kwargs: Any):
        self.kwargs = kwargs
        self._session = None  # Reuse session for multiple operations
        self._session_lock = asyncio.Lock()  # Protect session creation
        
        # Note: cleanup task will be started when needed
        # asyncio.create_task(async_cache.start_cleanup_task())
        
        super().__init__(**kwargs)
    
    async def start_cleanup_task(self):
        """Start the background cleanup task for the cache"""
        await async_cache.start_cleanup_task()
    
    async def _get_session(self) -> aiohttp.ClientSession:
        """Get or create a reusable session"""
        async with self._session_lock:
            if self._session is None or self._session.closed:
                self._session = aiohttp.ClientSession(
                    timeout=aiohttp.ClientTimeout(total=30),
                    connector=aiohttp.TCPConnector(limit=10, limit_per_host=5)
                )
            return self._session
    
    async def __aenter__(self):
        """Async context manager entry"""
        self._session = await self._get_session()
        return self
    
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        """Async context manager exit"""
        # Don't close session here - keep it for reuse
        pass
    
    async def close(self):
        """Close the session when done"""
        if self._session and not self._session.closed:
            await self._session.close()
    
    async def _download_file(self, url: str) -> str:
        """Download with session reuse"""
        temp_file = await async_cache.get_template(url, session=self._session)
        return temp_file
        
    async def _upload_file(self, file_content: bytes, filename: str, content_type: str,) -> tuple[str, str, str]:
        """Upload a file to a remote server and return the URL."""
        url, file_name, mime_type = await async_cache._do_upload(file_content, filename, content_type)
        
        return url, file_name, mime_type
    
    def _process_value(self, value: Any, value_type: str) -> Any:
        """Process value based on its type"""
        try:
            if value_type == "number":
                return float(value) if '.' in str(value) else int(value)
            elif value_type == "percentage":
                # Convert percentage to decimal (e.g., 5% -> 0.05)
                if isinstance(value, str) and value.endswith('%'):
                    return float(value.rstrip('%')) / 100
                return float(value) / 100 if float(value) > 1 else float(value)
            elif value_type == "boolean":
                return bool(value) if not isinstance(value, str) else value.lower() in ['true', '1', 'yes']
            elif value_type == "date":
                from datetime import datetime
                if isinstance(value, str):
                    # Try common date formats
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                        try:
                            return datetime.strptime(value, fmt)
                        except ValueError:
                            continue
                return value
            else:  # string
                return str(value)
        except Exception as e:
            return value  # Return original value as fallback
    
    async def calculate(
        self,
        file_url: str,
        dataRef: List[ExcelRef],
    ) -> str:
        """
        Perform calculations in an Excel file by updating specified cells and returning the updated file URL.
        Args:
            file_url (str): URL of the Excel file to download and modify.
            dataRef (List[ExcelRef]): List of cell references and their new values.
        Returns:
            str: File content in markdown format after calculations.
        """
        # Validation
        if not dataRef:
            raise ValueError("dataRef cannot be empty")
        
        if not file_url:
            raise ValueError("file_url cannot be empty")
        
        try:
            if not load_workbook:
                raise ImportError("openpyxl is not installed. Please install it to use ExcelCalculator.")
            
            # Download the cached template
            template_file = await self._download_file(file_url)
            
            # Create working copy with unique name to avoid conflicts
            working_file = f"./temps/working_{int(time.time() * 1000000)}_{os.getpid()}.xlsx"
            shutil.copy2(template_file, working_file)
            
            try:
                # Load the working copy
                wb = load_workbook(filename=working_file, data_only=False)
                ws = wb.active
                
                if not ws:
                    raise ValueError("The Excel file does not contain any worksheets.")
                
                # Update cells based on provided data
                for item in dataRef:
                    processed_value = self._process_value(item.value, item.value_type)
                    cell = ws[item.cell_ref]
                    # Check if cell is in a merged range
                    is_merged = any(cell.coordinate in merged_range for merged_range in ws.merged_cells.ranges)
                    if is_merged:
                        # Find the merge range and set value on the top-left cell
                        for merged_range in ws.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                ws[merged_range.start_cell.coordinate] = processed_value
                                break
                    else:
                        ws[item.cell_ref] = processed_value
                
                # Force recalculation of formulas
                wb.calculation.calcMode = 'auto'
                
                # Save to BytesIO
                from io import BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                # Process file after saving to Markdown
                from markitdown import MarkItDown
                
                md = MarkItDown()
                file_content = md.convert(output).markdown

                return file_content
                
            finally:
                # Always cleanup working file
                if os.path.exists(working_file):
                    os.remove(working_file)
        except FileNotFoundError as e:
            raise ValueError(f"Could not download file from URL: {file_url}")
        except Exception as e:
            return f"Error processing Excel file: {e}"


excel_calculator = ExcelCalculator()